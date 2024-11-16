import subprocess
import sys
from pathlib import Path
import os
import time
from openpyxl import load_workbook, Workbook
from tqdm import tqdm
import chromedriver_autoinstaller
from selenium import webdriver
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.options import Options

# Define paths
base_dir = Path.cwd()
venv_path = base_dir / '.venv'
excel_file_path = base_dir / "business_listings.xlsx"
status_file_path = base_dir / "status.txt"  # Path to status.txt in the base directory

# Constants for Excel sheet names
RESULTS_SHEET_NAME = "Results Sheet"
INPUT_SHEET_NAME = "Input Sheet"

def create_virtual_env():
    """Create a virtual environment if it does not exist."""
    if not venv_path.exists():
        subprocess.run([sys.executable, "-m", "venv", str(venv_path)], check=True)
        print("Virtual environment created successfully.")

def upgrade_pip():
    """Upgrade pip in the virtual environment."""
    pip_path = venv_path / 'Scripts' / 'python.exe' if os.name == 'nt' else venv_path / 'bin' / 'python'
    try:
        result = subprocess.run([str(pip_path), "-m", "pip", "install", "--upgrade", "pip"], check=True, capture_output=True)
        print(result.stdout.decode())
        print("Pip upgraded successfully.")
    except subprocess.CalledProcessError as e:
        print("Failed to upgrade pip. Proceeding with existing version.")

def install_required_packages():
    """Install required packages in the virtual environment."""
    pip_path = str(venv_path / 'Scripts' / 'pip' if os.name == 'nt' else venv_path / 'bin' / 'pip')
    dependencies = [
        "tqdm",
        "chromedriver-autoinstaller",
        "selenium",
        "openpyxl",
    ]
    try:
        subprocess.run([pip_path, "install"] + dependencies, check=True)
        print("Dependencies installed successfully.")
    except subprocess.CalledProcessError as e:
        print(f"Error installing dependencies: {e.stderr.decode()}")

def add_venv_to_sys_path():
    """Add the virtual environment's site-packages to sys.path."""
    site_packages_path = venv_path / 'Lib' / 'site-packages' if os.name == 'nt' else venv_path / 'lib' / f'python{sys.version_info.major}.{sys.version_info.minor}' / 'site-packages'
    if site_packages_path.exists():
        sys.path.insert(0, str(site_packages_path))
        print("Virtual environment packages added to sys.path.")
    else:
        print("site-packages path not found. Ensure virtual environment was created correctly.")

# Ensure required environment
create_virtual_env()
upgrade_pip()
install_required_packages()
add_venv_to_sys_path()

# Install the ChromeDriver version that matches the installed Chrome version
chromedriver_autoinstaller.install()

def setup_excel_file():
    """Create an Excel file with necessary sheets and headers if it does not exist."""
    if not excel_file_path.exists():
        print(f"Excel file does not exist. Creating file at {excel_file_path}")
        wb = Workbook()
        try:
            # Create Results Sheet
            results_sheet = wb.active
            results_sheet.title = RESULTS_SHEET_NAME
            headers = ["Rank", "Business Name", "Phone Number", "Business Page", "Website", 
                       "Category", "Rating", "Street Name", "Locality", "Region"]
            for col_num, header in enumerate(headers, 1):
                results_sheet[f"{get_column_letter(col_num)}1"] = header

            # Create Input Sheet
            input_sheet = wb.create_sheet(title=INPUT_SHEET_NAME)
            input_sheet["A1"] = "Search Terms"
            input_sheet["B1"] = "Geo Location Terms"
            input_sheet["C1"] = "Start Page"
            input_sheet["D1"] = "Max Pages"
            input_sheet["E1"] = "Run"
            input_sheet["E2"] = "Ready"

            wb.save(excel_file_path)
            print(f"Excel file created successfully at {excel_file_path}")
        except Exception as e:
            print(f"Error creating Excel file: {e}")
            raise
    else:
        print(f"Excel file already exists at {excel_file_path}")

def get_user_inputs():
    """Retrieve user inputs from the Input Sheet."""
    with load_workbook(excel_file_path) as wb:
        sheet = wb[INPUT_SHEET_NAME]
        run_status = sheet["E2"].value
        if run_status != "Run":
            sheet["E2"].value = "Ready"
            wb.save(excel_file_path)
            print("Run button not pressed. Exiting.")
            return None, None, None, None

        sheet["E2"].value = None
        wb.save(excel_file_path)

        search_terms = sheet["A2"].value
        geo_location_terms = sheet["B2"].value
        start_page = int(sheet["C2"].value)
        max_pages = int(sheet["D2"].value)
        return search_terms, geo_location_terms, start_page, max_pages

def update_status(status):
    """Update the Run status in the Excel file."""
    with load_workbook(excel_file_path) as wb:
        sheet = wb[INPUT_SHEET_NAME]
        sheet["E2"].value = status
        wb.save(excel_file_path)

def write_scraped_data(data):
    """Write scraped data to the Results Sheet."""
    with load_workbook(excel_file_path) as wb:
        results_sheet = wb[RESULTS_SHEET_NAME]
        for row_idx, entry in enumerate(data, start=2):
            for col_idx, (key, value) in enumerate(entry.items(), start=1):
                results_sheet[f"{get_column_letter(col_idx)}{row_idx}"] = value
        wb.save(excel_file_path)

class YellowPageScraper:
    BASE_URL = 'https://www.yellowpages.com'

    def __init__(self, search_terms, geo_location_terms, start_page, max_pages):
        self.search_terms = search_terms
        self.geo_location_terms = geo_location_terms
        self.current_page = start_page
        self.max_pages = max_pages
        self.chrome_options = Options()
        self.chrome_options.add_argument("--headless")
        self.driver = webdriver.Chrome(options=self.chrome_options)

    def scrape(self):
        """Main scraping function."""
        results = []
        try:
            while self.current_page <= self.max_pages:
                url = f"{self.BASE_URL}/search?search_terms={self.search_terms}&geo_location_terms={self.geo_location_terms}&page={self.current_page}"
                self.driver.get(url)
                time.sleep(6)
                cards = self.driver.find_elements("css selector", ".organic .srp-listing")
                if not cards:
                    break
                for card in tqdm(cards, desc=f"Scraping Page {self.current_page}", leave=False):
                    results.append(self.extract_business_listing(card))
                self.current_page += 1
        finally:
            self.driver.quit()
        return results

    def extract_business_listing(self, card):
        """Extract business info from a card."""
        # Your card parsing logic here
        return {}

# Main Execution
if __name__ == "__main__":
    try:
        setup_excel_file()
        search_terms, geo_location_terms, start_page, max_pages = get_user_inputs()
        if search_terms:
            update_status("Running")
            scraper = YellowPageScraper(search_terms, geo_location_terms, start_page, max_pages)
            results = scraper.scrape()
            write_scraped_data(results)
            update_status("Complete")
    except Exception as e:
        print(f"An error occurred: {e}")
        update_status("Error")
