import subprocess
import sys
from pathlib import Path
import os
import time


# Define paths
base_dir = Path.cwd()
venv_path = base_dir / '.venv'
excel_file_path = base_dir / "business_listings.xlsx"
status_file_path = base_dir / "status.txt"  # Path to status.txt in the base directory


# Constants for Excel sheet names
RESULTS_SHEET_NAME = "Results Sheet"
INPUT_SHEET_NAME = "Input Sheet"


def create_virtual_env():
    """Create a virtual environment if it does not exist."""
    if not venv_path.exists():
        subprocess.run([sys.executable, "-m", "venv", str(venv_path)], check=True)
        print("Virtual environment created successfully.")


def upgrade_pip():
    """Upgrade pip in the virtual environment."""
    pip_path = venv_path / 'Scripts' / 'python.exe' if os.name == 'nt' else venv_path / 'bin' / 'python'
    try:
        result = subprocess.run([str(pip_path), "-m", "pip", "install", "--upgrade", "pip"], check=True, capture_output=True)
        print(result.stdout.decode())
        print("Pip upgraded successfully.")
    except subprocess.CalledProcessError as e:
        print("Failed to upgrade pip. Proceeding with existing version.")


def install_required_packages():
    """Install required packages in the virtual environment."""
    pip_path = str(venv_path / 'Scripts' / 'pip' if os.name == 'nt' else venv_path / 'bin' / 'pip')
    dependencies = [
        "tqdm",
        "chromedriver-autoinstaller",
        "selenium",
        "openpyxl",
    ]
    try:
        subprocess.run([pip_path, "install"] + dependencies, check=True)
        print("Dependencies installed successfully.")
    except subprocess.CalledProcessError as e:
        print(f"Error installing dependencies: {e.stderr.decode()}")


def add_venv_to_sys_path():
    """Add the virtual environment's site-packages to sys.path."""
    site_packages_path = venv_path / 'Lib' / 'site-packages' if os.name == 'nt' else venv_path / 'lib' / f'python{sys.version_info.major}.{sys.version_info.minor}' / 'site-packages'
    if site_packages_path.exists():
        sys.path.insert(0, str(site_packages_path))
        print("Virtual environment packages added to sys.path.")
    else:
        print("site-packages path not found. Ensure virtual environment was created correctly.")



create_virtual_env()
upgrade_pip()
install_required_packages()
add_venv_to_sys_path()




try:
    from openpyxl import load_workbook, Workbook
    from tqdm import tqdm
    import chromedriver_autoinstaller
    from selenium import webdriver
    from openpyxl.utils import get_column_letter
    from selenium.webdriver.chrome.options import Options
    print("All packages imported successfully.")
except ImportError as e:
    print(f"Import error: {e}")



# Install the ChromeDriver version that matches the installed Chrome version
chromedriver_autoinstaller.install()



# Excel setup function
def setup_excel_file():
    """Create an Excel file with necessary sheets and headers if it does not exist."""
    if not excel_file_path.exists():
        print(f"Excel file does not exist. Creating file at {excel_file_path}")
        wb = Workbook()  # Create a new workbook
        try:
            # Create Results Sheet
            results_sheet = wb.active
            results_sheet.title = RESULTS_SHEET_NAME
            headers = ["Rank", "Business Name", "Phone Number", "Business Page", "Website", 
                       "Category", "Rating", "Street Name", "Locality", "Region"]
            for col_num, header in enumerate(headers, 1):
                results_sheet[f"{get_column_letter(col_num)}1"] = header

            # Create Input Sheet
            input_sheet = wb.create_sheet(title=INPUT_SHEET_NAME)
            input_sheet["A1"] = "Search Terms"
            input_sheet["B1"] = "Geo Location Terms"
            input_sheet["C1"] = "Start Page"
            input_sheet["D1"] = "Max Pages"
            input_sheet["E1"] = "Run"
            input_sheet["E2"] = "Ready"  # Default status

            # Save the workbook
            wb.save(excel_file_path)
            print(f"Excel file created successfully at {excel_file_path}")
        except Exception as e:
            print(f"Error creating Excel file: {e}")
            raise
    else:
        print(f"Excel file already exists at {excel_file_path}")



# Excel Interaction Functions
def get_user_inputs():
    """Retrieves user input from Excel sheet only if 'Run' button is pressed."""
    wb = load_workbook(excel_file_path)
    sheet = wb[INPUT_SHEET_NAME]
    run_status = sheet["E2"].value
    if run_status != "Run":
        sheet["E2"].value = "Ready"  # Reset status to Ready if not running
        wb.save(excel_file_path)
        wb.close()  # Close the workbook after saving
        print("Run button not pressed. Exiting.")
        return None, None, None, None

    # Clear the 'Run' status to avoid repeated runs
    sheet["E2"].value = None
    wb.save(excel_file_path)  # Save after clearing the Run status
    wb.close()  # Close the workbook after saving

    # Retrieve input values
    search_terms = sheet["A2"].value
    geo_location_terms = sheet["B2"].value
    start_page = int(sheet["C2"].value)
    max_pages = int(sheet["D2"].value)
    return search_terms, geo_location_terms, start_page, max_pages



def update_status(status):
    """Updates the Run status in the Excel file."""
    wb = load_workbook(excel_file_path)
    sheet = wb[INPUT_SHEET_NAME]
    sheet["E2"].value = status
    wb.save(excel_file_path)
    wb.close()  # Close the workbook after saving



def write_scraped_data(data):
    """Writes scraped data into the Results Sheet of the workbook."""
    wb = load_workbook(excel_file_path)  # Open workbook without 'with'
    try:
        results_sheet = wb[RESULTS_SHEET_NAME]

        # Clear existing data (except headers)
        for row in results_sheet.iter_rows(min_row=2, max_row=results_sheet.max_row):
            for cell in row:
                cell.value = None

        # Write new data
        for row_idx, entry in enumerate(data, start=2):
            for col_idx, (key, value) in enumerate(entry.items(), start=1):
                results_sheet[f"{get_column_letter(col_idx)}{row_idx}"] = value

        wb.save(excel_file_path)
        print("Results Sheet updated successfully.")
    except Exception as e:
        print(f"Error writing data to Results Sheet: {e}")
    finally:
        wb.close()  # Close the workbook




class YellowPageScraper:
    BASE_URL = 'https://www.yellowpages.com'

    def __init__(self, search_terms, geo_location_terms, start_page, max_pages):
        self.search_terms = search_terms
        self.geo_location_terms = geo_location_terms
        self.current_page = start_page
        self.max_pages = max_pages
        self.chrome_options = Options()
        self.chrome_options.add_argument("--headless")
        self.chrome_options.add_argument("--no-sandbox")
        self.chrome_options.add_argument("--disable-dev-shm-usage")
        self.driver = webdriver.Chrome(options=self.chrome_options)

        # Load the Excel workbook
        try:
            self.wb = load_workbook(excel_file_path)  # Load the workbook
            self.results_sheet = self.wb[RESULTS_SHEET_NAME]  # Access the Results Sheet
            self.clear_results_sheet()  # Clear any existing data in the Results Sheet
        except Exception as e:
            print(f"Error loading the workbook: {e}")

    def clear_results_sheet(self):
        """Clears existing data in the Results Sheet."""
        for row in self.results_sheet.iter_rows(min_row=2, max_row=self.results_sheet.max_row):
            for cell in row:
                cell.value = None

    def extract_business_listing(self, card):
        """Extracts business information from a card element."""
        rank = card.find_element("css selector", ".info-primary h2").text.strip() if card.find_elements("css selector", ".info-primary h2") else ''
        business_name = card.find_element("css selector", ".business-name span").text.strip() if card.find_elements("css selector", ".business-name span") else ''
        phone_number = card.find_element("css selector", ".phones").text.strip() if card.find_elements("css selector", ".phones") else ''
        business_page = card.find_element("css selector", ".business-name").get_attribute('href') if card.find_elements("css selector", ".business-name") else ''
        website = card.find_element("css selector", ".track-visit-website").get_attribute('href') if card.find_elements("css selector", ".track-visit-website") else ''
        category = ', '.join([a.text.strip() for a in card.find_elements("css selector", ".categories a")]) if card.find_elements("css selector", ".categories a") else ''
        rating = card.find_element("css selector", ".ratings .count").text.strip('()') if card.find_elements("css selector", ".ratings .count") else ''
        street_name = card.find_element("css selector", ".street-address").text.strip() if card.find_elements("css selector", ".street-address") else ''
        locality = card.find_element("css selector", ".locality").text.strip() if card.find_elements("css selector", ".locality") else ''
        locality_parts = locality.split(",") if locality else ['', '']
        locality, region = locality_parts[0].strip(), locality_parts[1].strip() if len(locality_parts) > 1 else ''
        return {
            "Rank": rank,
            "Business Name": business_name,
            "Phone Number": phone_number,
            "Business Page": business_page,
            "Website": website,
            "Category": category,
            "Rating": rating,
            "Street Name": street_name,
            "Locality": locality,
            "Region": region
        }

    def scrape(self):
        """Main scraping function with a maximum page limit."""
        results = []
        try:
            while self.current_page <= self.max_pages:
                url = f"{self.BASE_URL}/search?search_terms={self.search_terms}&geo_location_terms={self.geo_location_terms}&page={self.current_page}"
                self.driver.get(url)
                time.sleep(6)
                cards = self.driver.find_elements("css selector", ".organic .srp-listing")
                if not cards:
                    print(f"No more cards found on page {self.current_page}.")
                    break

                for card in tqdm(cards, desc=f"Scraping Listings (Page {self.current_page}):", leave=False):
                    business_info = self.extract_business_listing(card)
                    results.append(business_info)
                self.current_page += 1
                print(f"Page {self.current_page} scraped successfully.")
        except Exception as e:
            print(f"Error during scraping: {e}")
        finally:
            self.driver.quit()
        return results


import traceback

# Main Execution Block
if __name__ == "__main__":
    status_file_path = "status.txt"  # Path to the status file

    try:
        # Ensure the Excel file is set up properly
        setup_excel_file()
        
        # Get user inputs
        search_terms, geo_location_terms, start_page, max_pages = get_user_inputs()
        
        # Validate inputs
        if search_terms and geo_location_terms and start_page is not None and max_pages is not None:
            # Initialize and run the scraper
            scraper = YellowPageScraper(search_terms, geo_location_terms, start_page, max_pages)
            update_status("Running")
            results = scraper.scrape()  # Assuming this returns the scraped data
            write_scraped_data(results)  # Write the scraped data to Excel
            update_status("Complete")
            print("Scraping completed successfully.")
        else:
            error_message = "Exiting due to invalid inputs."
            print(error_message)
            update_status("Error")
            raise ValueError(error_message)

    except FileNotFoundError as fnfe:
        error_message = f"File not found error: {fnfe}"
        print(error_message)
        update_status("Error")
        # Log the detailed error to the status file
        with open(status_file_path, "w") as status_file:
            status_file.write(f"Error: {error_message}\n{traceback.format_exc()}")
    
    except Exception as e:
        error_message = f"An unexpected error occurred: {e}"
        print(error_message)
        update_status("Error")
        # Log the detailed error to the status file
        with open(status_file_path, "w") as status_file:
            status_file.write(f"Error: {error_message}\n{traceback.format_exc()}")
    
    else:
        # If no exceptions, write 'Complete' to the status file
        with open(status_file_path, "w") as status_file:
            status_file.write("Complete")
    
    finally:
        print("Execution finished. Check status.txt for details.")